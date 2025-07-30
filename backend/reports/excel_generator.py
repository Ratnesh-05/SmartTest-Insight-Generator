"""
Excel Generator Module
Handles Excel report generation using OpenPyXL
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Optional
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

class ExcelReportGenerator:
    """Generates comprehensive Excel reports for performance testing"""
    
    def __init__(self):
        self.workbook = None
        self.styles = self._create_styles()
        
    def _create_styles(self) -> Dict:
        """Create Excel styles for formatting"""
        return {
            'header': {
                'font': Font(bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'subheader': {
                'font': Font(bold=True, color='000000'),
                'fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'data': {
                'font': Font(color='000000'),
                'alignment': Alignment(horizontal='left', vertical='center')
            },
            'number': {
                'font': Font(color='000000'),
                'alignment': Alignment(horizontal='right', vertical='center')
            },
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def create_performance_report(self, data: Dict, output_path: str) -> str:
        """
        Create comprehensive performance report in Excel
        
        Args:
            data: Performance data dictionary
            output_path: Path to save the Excel file
            
        Returns:
            str: Path to created Excel file
        """
        try:
            self.workbook = Workbook()
            
            # Remove default sheet
            if 'Sheet' in self.workbook.sheetnames:
                self.workbook.remove(self.workbook['Sheet'])
            
            # Create different sheets
            self._create_summary_sheet(data)
            self._create_metrics_sheet(data)
            self._create_trends_sheet(data)
            self._create_anomalies_sheet(data)
            self._create_recommendations_sheet(data)
            
            # Save workbook
            self.workbook.save(output_path)
            logger.info(f"Excel report created successfully at {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error creating Excel report: {str(e)}")
            raise
    
    def _create_summary_sheet(self, data: Dict) -> None:
        """Create executive summary sheet"""
        ws = self.workbook.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = "Performance Test Report - Executive Summary"
        ws.merge_cells('A1:F1')
        self._apply_style(ws['A1'], 'header')
        
        # Report info
        row = 3
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Key metrics summary
        row += 2
        ws[f'A{row}'] = "Key Performance Metrics"
        self._apply_style(ws[f'A{row}'], 'subheader')
        
        if 'metrics' in data:
            metrics = data['metrics']
            row += 2
            
            # Response time metrics
            if 'response_time' in metrics:
                rt = metrics['response_time']
                ws[f'A{row}'] = "Average Response Time (ms):"
                ws[f'B{row}'] = round(rt.get('mean', 0), 2)
                row += 1
                ws[f'A{row}'] = "95th Percentile (ms):"
                ws[f'B{row}'] = round(rt.get('p95', 0), 2)
                row += 1
                ws[f'A{row}'] = "99th Percentile (ms):"
                ws[f'B{row}'] = round(rt.get('p99', 0), 2)
                row += 1
            
            # Error metrics
            if 'errors' in metrics:
                errors = metrics['errors']
                ws[f'A{row}'] = "Error Rate (%):"
                ws[f'B{row}'] = round(errors.get('error_rate', 0), 2)
                row += 1
                ws[f'A{row}'] = "Total Errors:"
                ws[f'B{row}'] = errors.get('error_requests', 0)
                row += 1
        
        # Test summary
        if 'summary' in data:
            summary = data['summary']
            row += 2
            ws[f'A{row}'] = "Test Summary"
            self._apply_style(ws[f'A{row}'], 'subheader')
            row += 1
            
            ws[f'A{row}'] = "Total Records:"
            ws[f'B{row}'] = summary.get('total_entries', 0)
            row += 1
            
            if 'time_range' in summary:
                time_range = summary['time_range']
                ws[f'A{row}'] = "Test Duration:"
                ws[f'B{row}'] = f"{time_range.get('duration_hours', 0):.2f} hours"
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_metrics_sheet(self, data: Dict) -> None:
        """Create detailed metrics sheet"""
        ws = self.workbook.create_sheet("Detailed Metrics")
        
        # Title
        ws['A1'] = "Detailed Performance Metrics"
        ws.merge_cells('A1:D1')
        self._apply_style(ws['A1'], 'header')
        
        if 'metrics' not in data:
            return
            
        metrics = data['metrics']
        row = 3
        
        # Response time metrics
        if 'response_time' in metrics:
            rt = metrics['response_time']
            ws[f'A{row}'] = "Response Time Metrics"
            self._apply_style(ws[f'A{row}'], 'subheader')
            row += 1
            
            # Headers
            ws[f'A{row}'] = "Metric"
            ws[f'B{row}'] = "Value (ms)"
            self._apply_style(ws[f'A{row}'], 'subheader')
            self._apply_style(ws[f'B{row}'], 'subheader')
            row += 1
            
            # Data
            for metric, value in rt.items():
                ws[f'A{row}'] = metric.replace('_', ' ').title()
                ws[f'B{row}'] = round(value, 2)
                row += 1
            
            row += 1
        
        # Throughput metrics
        if 'throughput' in metrics:
            tp = metrics['throughput']
            ws[f'A{row}'] = "Throughput Metrics"
            self._apply_style(ws[f'A{row}'], 'subheader')
            row += 1
            
            ws[f'A{row}'] = "Metric"
            ws[f'B{row}'] = "Value"
            self._apply_style(ws[f'A{row}'], 'subheader')
            self._apply_style(ws[f'B{row}'], 'subheader')
            row += 1
            
            for metric, value in tp.items():
                ws[f'A{row}'] = metric.replace('_', ' ').title()
                ws[f'B{row}'] = value
                row += 1
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_trends_sheet(self, data: Dict) -> None:
        """Create trends analysis sheet"""
        ws = self.workbook.create_sheet("Trends Analysis")
        
        ws['A1'] = "Performance Trends Analysis"
        ws.merge_cells('A1:E1')
        self._apply_style(ws['A1'], 'header')
        
        if 'trends' not in data:
            return
            
        trends = data['trends']
        row = 3
        
        # Response time trends
        if 'response_time_trend' in trends:
            rt_trend = trends['response_time_trend']
            ws[f'A{row}'] = "Response Time Trends"
            self._apply_style(ws[f'A{row}'], 'subheader')
            row += 1
            
            if 'hourly_pattern' in rt_trend:
                ws[f'A{row}'] = "Hourly Pattern"
                self._apply_style(ws[f'A{row}'], 'subheader')
                row += 1
                
                ws[f'A{row}'] = "Hour"
                ws[f'B{row}'] = "Avg Response Time (ms)"
                self._apply_style(ws[f'A{row}'], 'subheader')
                self._apply_style(ws[f'B{row}'], 'subheader')
                row += 1
                
                for hour, avg_time in rt_trend['hourly_pattern'].items():
                    ws[f'A{row}'] = hour
                    ws[f'B{row}'] = round(avg_time, 2)
                    row += 1
                
                row += 1
    
    def _create_anomalies_sheet(self, data: Dict) -> None:
        """Create anomalies analysis sheet"""
        ws = self.workbook.create_sheet("Anomalies")
        
        ws['A1'] = "Performance Anomalies"
        ws.merge_cells('A1:F1')
        self._apply_style(ws['A1'], 'header')
        
        if 'anomalies' not in data:
            return
            
        anomalies = data['anomalies']
        row = 3
        
        # Anomaly summary
        ws[f'A{row}'] = f"Total Anomalies Detected: {len(anomalies)}"
        self._apply_style(ws[f'A{row}'], 'subheader')
        row += 2
        
        if anomalies:
            # Headers
            ws[f'A{row}'] = "Index"
            ws[f'B{row}'] = "Value"
            ws[f'C{row}'] = "Z-Score"
            ws[f'D{row}'] = "Timestamp"
            
            for col in ['A', 'B', 'C', 'D']:
                self._apply_style(ws[f'{col}{row}'], 'subheader')
            
            row += 1
            
            # Data
            for anomaly in anomalies:
                ws[f'A{row}'] = anomaly.get('index', '')
                ws[f'B{row}'] = round(anomaly.get('value', 0), 2)
                ws[f'C{row}'] = round(anomaly.get('z_score', 0), 2)
                ws[f'D{row}'] = anomaly.get('timestamp', '')
                row += 1
    
    def _create_recommendations_sheet(self, data: Dict) -> None:
        """Create recommendations sheet"""
        ws = self.workbook.create_sheet("Recommendations")
        
        ws['A1'] = "Performance Recommendations"
        ws.merge_cells('A1:D1')
        self._apply_style(ws['A1'], 'header')
        
        row = 3
        
        # General recommendations
        recommendations = [
            "Monitor response time trends regularly",
            "Set up alerting for performance degradation",
            "Implement caching for frequently accessed data",
            "Optimize database queries",
            "Consider load balancing for high traffic",
            "Regular performance testing in CI/CD pipeline"
        ]
        
        if 'insights' in data and 'recommendations' in data['insights']:
            recommendations.extend(data['insights']['recommendations'])
        
        ws[f'A{row}'] = "Recommendations"
        self._apply_style(ws[f'A{row}'], 'subheader')
        row += 1
        
        for i, rec in enumerate(recommendations, 1):
            ws[f'A{row}'] = f"{i}."
            ws[f'B{row}'] = rec
            row += 1
        
        # Risk areas
        if 'insights' in data and 'risk_areas' in data['insights']:
            row += 2
            ws[f'A{row}'] = "Risk Areas"
            self._apply_style(ws[f'A{row}'], 'subheader')
            row += 1
            
            for i, risk in enumerate(data['insights']['risk_areas'], 1):
                ws[f'A{row}'] = f"{i}."
                ws[f'B{row}'] = risk
                row += 1
    
    def _apply_style(self, cell, style_name: str) -> None:
        """Apply predefined style to a cell"""
        if style_name in self.styles:
            style = self.styles[style_name]
            if 'font' in style:
                cell.font = style['font']
            if 'fill' in style:
                cell.fill = style['fill']
            if 'alignment' in style:
                cell.alignment = style['alignment']
            if 'border' in style:
                cell.border = style['border']
    
    def create_raw_data_sheet(self, df: pd.DataFrame, sheet_name: str = "Raw Data") -> None:
        """
        Create sheet with raw performance data
        
        Args:
            df: DataFrame with raw data
            sheet_name: Name of the sheet
        """
        if self.workbook is None:
            self.workbook = Workbook()
            
        ws = self.workbook.create_sheet(sheet_name)
        
        # Add dataframe to sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style header row
        for cell in ws[1]:
            self._apply_style(cell, 'header')
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width